# -*- coding: utf-8 -*-
"""下拉选项模板接口：CRUD + Excel 上传解析。"""
import json
from io import BytesIO

from flask import Blueprint, request, jsonify

from ..db import get_db, gen_id, now_str, row_to_dict

bp = Blueprint('option_sets', __name__)


def _os_to_dict(row):
    d = row_to_dict(row)
    if d and 'options_json' in d:
        try:
            d['options'] = json.loads(d['options_json'])
        except Exception:
            d['options'] = []
        d.pop('options_json', None)
    return d


@bp.route('/api/option-sets', methods=['GET'])
def api_list_option_sets():
    db = get_db()
    rows = db.execute('SELECT * FROM option_sets ORDER BY created_at DESC').fetchall()
    result = []
    for r in rows:
        d = _os_to_dict(r)
        d['count'] = len(d.get('options', []))
        result.append(d)
    return jsonify({'code': 0, 'data': result})


@bp.route('/api/option-sets', methods=['POST'])
def api_create_option_set():
    """创建选项模板。支持 JSON {name, options:[...]} 或 Excel 上传。"""
    # Excel 上传模式
    if 'file' in request.files:
        f = request.files['file']
        if not f.filename:
            return jsonify({'code': 1, 'message': '文件名为空'}), 400
        file_bytes = f.read()
        try:
            from openpyxl import load_workbook
            wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
            ws = wb.active
            options = []
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None and str(cell).strip():
                        options.append(str(cell).strip())
        except Exception as e:
            return jsonify({'code': 1, 'message': f'解析失败: {e}'}), 500
        name = request.form.get('name', f.filename.rsplit('.', 1)[0])
    else:
        data = request.get_json() or {}
        name = (data.get('name') or '').strip()
        options = data.get('options') or []
        if isinstance(options, str):
            options = [o.strip() for o in options.split('\n') if o.strip()]

    if not name:
        return jsonify({'code': 1, 'message': '名称不能为空'}), 400
    if not options:
        return jsonify({'code': 1, 'message': '选项不能为空'}), 400

    # 去重保序
    seen = set()
    unique = []
    for o in options:
        if o not in seen:
            seen.add(o)
            unique.append(o)

    oid = gen_id()
    db = get_db()
    db.execute('INSERT INTO option_sets (id, name, options_json, created_at) VALUES (?, ?, ?, ?)',
               (oid, name, json.dumps(unique, ensure_ascii=False), now_str()))
    db.commit()
    row = db.execute('SELECT * FROM option_sets WHERE id = ?', (oid,)).fetchone()
    return jsonify({'code': 0, 'data': _os_to_dict(row)})


@bp.route('/api/option-sets/<oid>', methods=['PUT'])
def api_update_option_set(oid):
    data = request.get_json() or {}
    db = get_db()
    row = db.execute('SELECT * FROM option_sets WHERE id = ?', (oid,)).fetchone()
    if not row:
        return jsonify({'code': 1, 'message': '选项模板不存在'}), 404
    name = (data.get('name') or '').strip()
    options = data.get('options') or []
    if isinstance(options, str):
        options = [o.strip() for o in options.split('\n') if o.strip()]
    if name:
        db.execute('UPDATE option_sets SET name = ? WHERE id = ?', (name, oid))
    if options:
        seen = set()
        unique = []
        for o in options:
            if o not in seen:
                seen.add(o)
                unique.append(o)
        db.execute('UPDATE option_sets SET options_json = ? WHERE id = ?',
                   (json.dumps(unique, ensure_ascii=False), oid))
    db.commit()
    row = db.execute('SELECT * FROM option_sets WHERE id = ?', (oid,)).fetchone()
    return jsonify({'code': 0, 'data': _os_to_dict(row)})


@bp.route('/api/option-sets/<oid>', methods=['DELETE'])
def api_delete_option_set(oid):
    db = get_db()
    db.execute('DELETE FROM option_sets WHERE id = ?', (oid,))
    db.commit()
    return jsonify({'code': 0, 'data': True})
