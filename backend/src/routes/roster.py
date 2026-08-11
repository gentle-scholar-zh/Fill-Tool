# -*- coding: utf-8 -*-
"""名单接口：列表、上传解析、详情、删除、验证。"""
import os
import json
from io import BytesIO

from flask import Blueprint, request, jsonify

from ..config import RST_DIR
from ..db import get_db, gen_id, now_str, row_to_dict
from ..utils import (detect_header_row, auto_detect_identity_fields,
                    deduplicate_roster, safe_filename)

bp = Blueprint('roster', __name__)


@bp.route('/api/roster', methods=['GET'])
def api_get_roster():
    db = get_db()
    rows = db.execute('SELECT * FROM rosters ORDER BY created_at DESC').fetchall()
    result = []
    for r in rows:
        d = row_to_dict(r)
        d.pop('rows_json', None)
        try:
            d['headers'] = json.loads(d.pop('headers_json', '[]'))
        except Exception:
            d['headers'] = []
        result.append(d)
    return jsonify({'code': 0, 'data': result})


@bp.route('/api/roster/upload', methods=['POST'])
def api_upload_roster():
    """上传名单 Excel，解析并存储。"""
    if 'file' not in request.files:
        return jsonify({'code': 1, 'message': '未提供文件'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'code': 1, 'message': '文件名为空'}), 400
    file_bytes = f.read()

    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return jsonify({'code': 1, 'message': '文件为空'}), 400
        header_idx, headers, data_rows = detect_header_row(all_rows)
    except Exception as e:
        return jsonify({'code': 1, 'message': f'解析失败: {e}'}), 500

    rid = gen_id()
    roster_name = request.form.get('name', f.filename.rsplit('.', 1)[0])
    category_id = request.form.get('category_id', '')
    file_path = os.path.join(RST_DIR, f'{rid}_{safe_filename(f.filename)}')
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    with open(file_path, 'wb') as fp:
        fp.write(file_bytes)

    id_field, name_field = auto_detect_identity_fields(headers)
    data_rows, original_count = deduplicate_roster(data_rows, headers, id_field, name_field)
    dedup_removed = original_count - len(data_rows)

    db = get_db()
    db.execute('''INSERT INTO rosters (id, name, file_name, file_path, total, headers_json, rows_json, category_id, created_at)
                  VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''',
               (rid, roster_name, f.filename, file_path, len(data_rows),
                json.dumps(headers, ensure_ascii=False),
                json.dumps(data_rows, ensure_ascii=False), category_id or None, now_str()))
    if category_id:
        try:
            db.execute('UPDATE categories SET count = count + 1 WHERE id = ?', (category_id,))
        except Exception:
            pass
    db.commit()
    return jsonify({'code': 0, 'data': {
        'id': rid,
        'name': roster_name,
        'fileName': f.filename,
        'total': len(data_rows),
        'originalCount': original_count,
        'dedupRemoved': dedup_removed,
        'headers': headers,
        'allFields': headers,
        'preview': data_rows[:10],
        'id_field': id_field,
        'name_field': name_field,
        'header_row': header_idx + 1,
        'category_id': category_id or None,
    }})


@bp.route('/api/roster/<rid>', methods=['GET'])
def api_get_roster_detail(rid):
    """获取单个名单详情（含完整数据）。"""
    db = get_db()
    row = db.execute('SELECT * FROM rosters WHERE id = ?', (rid,)).fetchone()
    if not row:
        return jsonify({'code': 1, 'message': '名单不存在'}), 404
    d = row_to_dict(row)
    try:
        d['headers'] = json.loads(d.pop('headers_json', '[]'))
        d['rows'] = json.loads(d.pop('rows_json', '[]'))
    except Exception:
        d['headers'], d['rows'] = [], []
    d.pop('file_path', None)
    return jsonify({'code': 0, 'data': d})


@bp.route('/api/roster/<rid>', methods=['PUT'])
def api_update_roster(rid):
    """更新名单所属分组（category_id）。"""
    data = request.get_json() or {}
    db = get_db()
    row = db.execute('SELECT * FROM rosters WHERE id = ?', (rid,)).fetchone()
    if not row:
        return jsonify({'code': 1, 'message': '名单不存在'}), 404
    cat_id = data.get('category_id', '')
    # 更新分类计数
    old_cat = row['category_id']
    if old_cat != (cat_id or None):
        if old_cat:
            db.execute('UPDATE categories SET count = count - 1 WHERE id = ? AND count > 0', (old_cat,))
        if cat_id:
            db.execute('UPDATE categories SET count = count + 1 WHERE id = ?', (cat_id,))
    db.execute('UPDATE rosters SET category_id = ? WHERE id = ?', (cat_id or None, rid))
    db.commit()
    return jsonify({'code': 0, 'data': True})


@bp.route('/api/roster/<rid>', methods=['DELETE'])
def api_delete_roster(rid):
    db = get_db()
    row = db.execute('SELECT * FROM rosters WHERE id = ?', (rid,)).fetchone()
    if not row:
        return jsonify({'code': 1, 'message': '名单不存在'}), 404
    try:
        if os.path.exists(row['file_path']):
            os.remove(row['file_path'])
    except Exception:
        pass
    db.execute('DELETE FROM template_roster WHERE roster_id = ?', (rid,))
    db.execute('DELETE FROM rosters WHERE id = ?', (rid,))
    db.commit()
    return jsonify({'code': 0, 'data': True})


@bp.route('/api/roster/verify', methods=['POST'])
def api_verify_roster():
    """验证用户提交的学号+姓名，并返回该学生名下所有名额（每行一个独立名额，支持重复姓名多次提交）。"""
    data = request.get_json() or {}
    tid = data.get('template_id')
    student_id = (data.get('student_id') or '').strip()
    student_name = (data.get('student_name') or '').strip()
    if not tid or not student_id or not student_name:
        return jsonify({'code': 1, 'message': '请填写学号和姓名'}), 400

    db = get_db()
    link = db.execute('SELECT * FROM template_roster WHERE template_id = ?', (tid,)).fetchone()
    if not link:
        return jsonify({'code': 1, 'message': '该模板未关联名单，无需验证'}), 400
    roster = db.execute('SELECT * FROM rosters WHERE id = ?', (link['roster_id'],)).fetchone()
    if not roster:
        return jsonify({'code': 1, 'message': '名单不存在'}), 404

    rows = json.loads(roster['rows_json'])
    id_field = link['id_field']
    name_field = link['name_field']

    award_field = link['award_field'] or ''

    # 该学生名下所有名额行（保留重复姓名 -> 多个名额）
    matched_rows = [r for r in rows
                    if str(r.get(id_field, '')).strip() == student_id
                    and str(r.get(name_field, '')).strip() == student_name]
    if not matched_rows:
        return jsonify({'code': 1, 'message': '学号或姓名不匹配，请确认后重试'})

    # 每行名额的提交状态
    slots = []
    for r in matched_rows:
        rid = r.get('row_id') or ''
        sub = db.execute('SELECT * FROM submissions WHERE template_id = ? AND roster_row_id = ?',
                         (tid, rid)).fetchone()
        slot = {
            'row_id': rid,
            'roster_data': r,
            'submitted': sub is not None,
            'sub_id': sub['id'] if sub else None,
            'version': sub['version'] if sub else 0,
            'edit_count': sub['edit_count'] if sub else 0,
            'data': json.loads(sub['data_json']) if sub else None,
        }
        slots.append(slot)

    filled = sum(1 for s in slots if s['submitted'])
    return jsonify({'code': 0, 'data': {
        'verified': True,
        'student_id': student_id,
        'student_name': student_name,
        'id_field': id_field,
        'name_field': name_field,
        'award_field': award_field,
        'slot_count': len(slots),
        'filled_count': filled,
        'slots': slots,
    }})



@bp.route('/api/roster/dispatch', methods=['POST'])
def api_dispatch_roster():
    """模拟发送通知（预留）。"""
    return jsonify({'code': 0, 'data': {'sent': 0}})
