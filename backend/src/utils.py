# -*- coding: utf-8 -*-
"""通用工具：字段类型推断、文档占位符提取、表头检测、名单去重、ORM 序列化。"""
import re
import json
import zipfile
from io import BytesIO

from .db import row_to_dict
from .config import TPL_DIR, SUB_DIR


# ============================ 字段类型推断 ============================
def infer_type(name: str) -> dict:
    """根据字段名推断类型，特殊字段加 pattern 校验。前端 helpers.js 保持同步。"""
    name = (name or '').strip()
    exact_map = {
        '性别':    {'type': 'select', 'options': ['男', '女']},
        'sex':     {'type': 'select', 'options': ['男', '女']},
        '生日':    {'type': 'date'},
        '出生日期': {'type': 'date'},
        'birthday': {'type': 'date'},
    }
    for k, v in exact_map.items():
        if name == k:
            return v
    if any(k in name for k in ['手机号', '电话', 'mobile', 'phone']):
        return {'type': 'text', 'pattern': r'^1\d{10}$', 'placeholder': '请输入11位手机号', 'hint': '11位手机号'}
    if any(k in name for k in ['身份证', 'id card', 'idcard']):
        return {'type': 'text', 'pattern': r'^\d{17}[\dXx]$', 'placeholder': '请输入18位身份证号', 'hint': '18位身份证号'}
    if any(k in name for k in ['邮箱', 'email', 'mail']):
        return {'type': 'text', 'pattern': r'^[\w.+-]+@[\w-]+\.[\w.-]+$', 'placeholder': 'name@example.com', 'hint': '邮箱格式'}
    if any(k in name for k in ['学号', '工号', '编号', 'no', 'id']):
        return {'type': 'text', 'pattern': r'^\w{4,20}$', 'placeholder': '请输入编号', 'hint': '4-20位字母数字'}
    if any(k in name for k in ['年龄', 'age']):
        return {'type': 'number', 'pattern': r'^\d{1,3}$', 'placeholder': '请输入年龄', 'hint': '1-3位数字'}
    if any(k in name for k in ['金额', '价格', 'price', 'amount', 'money']):
        return {'type': 'number', 'pattern': r'^\d+(\.\d{1,2})?$', 'placeholder': '0.00', 'hint': '最多两位小数'}
    if any(k in name for k in ['班级', 'class', '年级', 'grade', '部门', 'department']):
        return {'type': 'text', 'placeholder': '请输入'}
    if any(k in name for k in ['理由', '说明', '描述', 'remark', 'reason', 'description']):
        return {'type': 'textarea', 'placeholder': '请输入详细说明'}
    return {'type': 'text'}


def docx_to_type(inferred_type: str) -> str:
    """将推断类型映射为前端展示的友好类型名。"""
    return {
        'text': '单行文本',
        'textarea': '多行文本',
        'number': '数字',
        'date': '日期选择',
        'select': '下拉选择',
    }.get(inferred_type, '单行文本')


# ============================ 占位符提取 ============================
def extract_placeholders(file_bytes: bytes) -> list:
    """从 docx 提取所有 {{字段名}}，清理空格并去重保序。"""
    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as zf:
            xml = zf.read('word/document.xml').decode('utf-8', errors='ignore')
        texts = re.findall(r'<w:t[^>]*>(.*?)</w:t>', xml, re.DOTALL)
        full_text = ''.join(texts)
        matches = re.findall(r'\{\{\s*([^{}]+?)\s*\}\}', full_text)
        seen, ordered = set(), []
        for m in matches:
            m = m.strip()
            if m and m not in seen:
                seen.add(m)
                ordered.append(m)
        return ordered
    except Exception as e:
        print('提取占位符出错:', e)
        return []


def extract_placeholders_xlsx(file_bytes: bytes) -> list:
    """从 xlsx 提取第 1 行作为字段名（备用）。"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        row1 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not row1:
            return []
        out = []
        for v in row1:
            if v is None:
                continue
            s = str(v).strip()
            if s and s not in out:
                out.append(s)
        return out
    except Exception as e:
        print('提取 xlsx 表头出错:', e)
        return []


def safe_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|\r\n]', '_', str(name or '未知')).strip() or '未知'


# ============================ 表头自动检测 ============================
_HEADER_KEYWORDS = {
    '学号', '姓名', '序号', '编号', '名字', '班级', '年级', '身份证', '手机',
    '电话', '邮箱', '性别', '年龄', '院系', '专业', '系别', '部门', '职务',
    '卡号', '金额', '成绩', '排名', '奖项', '备注', '地址', '生日', '日期',
    '考号', '工号', '准考证号', '持卡人', '班级人数',
}


def detect_header_row(all_rows, max_scan=15):
    """扫描前 max_scan 行，找到含关键词最多的行作为表头。返回 (header_index, headers, data_rows)。"""
    best_idx, best_score = 0, 0
    for i, row in enumerate(all_rows[:max_scan]):
        score = 0
        for cell in row:
            if cell is None:
                continue
            cell_str = str(cell).strip().replace(' ', '').replace('　', '')
            if cell_str in _HEADER_KEYWORDS:
                score += 2
            elif any(kw in cell_str for kw in _HEADER_KEYWORDS):
                score += 1
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 2:
            score += 1
        if score > best_score:
            best_score = score
            best_idx = i

    headers = [str(v).strip() if v is not None else '' for v in all_rows[best_idx]]
    for i, h in enumerate(headers):
        if not h:
            headers[i] = f'列{i+1}'

    data_rows = []
    for row in all_rows[best_idx + 1:]:
        if all(v is None or str(v).strip() == '' for v in row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            row_dict[h] = str(row[i]).strip() if i < len(row) and row[i] is not None else ''
        data_rows.append(row_dict)
    return best_idx, headers, data_rows


# ============================ 名单身份识别与去重 ============================
def auto_detect_identity_fields(headers):
    """根据表头自动识别学号、姓名字段，支持多种常见命名。"""
    if not headers:
        return '学号', '姓名'
    id_patterns = ['学号', '编号', '学生编号', '学员编号', '工号', '考号', '准考证号', '身份证号']
    name_patterns = ['姓名', '名字', '学生姓名', '真实姓名', '考生姓名']

    id_field = name_field = None
    for h in headers:
        h_clean = str(h).strip().replace(' ', '').replace('　', '')
        if not id_field:
            for kw in id_patterns:
                if h_clean == kw or (kw in h_clean and len(h_clean) <= len(kw) + 2):
                    id_field = h
                    break
        if not name_field:
            for kw in name_patterns:
                if h_clean == kw:
                    name_field = h
                    break
    if not id_field and len(headers) >= 2:
        id_field = headers[1]
    if not name_field and headers:
        name_field = headers[0]
    return id_field or '学号', name_field or '姓名'


def deduplicate_roster(data_rows, headers, id_field, name_field):
    """名单处理：每一行视为一个独立名额（例如同一人获一/二等奖应分别填写）。

    不再按学号/姓名合并——保留重复姓名。仅移除完全空白行和逐字段完全相同的
    完全重复行（同一人同一奖项被误复制两次的情况）。为每行分配稳定的 row_id，
    用于后续按名额定位提交记录。
    """
    if not data_rows:
        return [], 0
    original_count = len(data_rows)

    seen_full = set()
    slots = []
    for idx, row in enumerate(data_rows, start=1):
        # 跳过全空行
        if all(not str(v or '').strip() for v in row.values()):
            continue
        # 完全重复行识别（逐字段文本一致）
        sig = tuple(str(row.get(h, '')).strip() for h in headers)
        if sig in seen_full:
            continue
        seen_full.add(sig)
        r = dict(row)
        r['row_id'] = f'r{idx:04d}'
        slots.append(r)

    seq_field = next((h for h in headers if h in ('序号', '编号', 'NO', 'No', 'no', '序', '号')), None)
    if seq_field:
        def sort_key(r):
            val = r.get(seq_field, '')
            try:
                return int(float(val))
            except Exception:
                return 999999
        slots.sort(key=sort_key)
        # 重排 row_id 保持与可见顺序一致
        for i, r in enumerate(slots, start=1):
            r['row_id'] = f'r{i:04d}'
    return slots, original_count


# ============================ ORM 序列化 ============================
def tpl_to_dict(row):
    d = row_to_dict(row)
    if d and 'fields_json' in d:
        try:
            d['fields'] = json.loads(d['fields_json'])
        except Exception:
            d['fields'] = []
        d.pop('fields_json', None)
        d.pop('file_path', None)
    return d


def sub_to_dict(row):
    d = row_to_dict(row)
    if d and 'data_json' in d:
        try:
            d['data'] = json.loads(d['data_json'])
        except Exception:
            d['data'] = {}
        d.pop('data_json', None)
    return d


def build_filename_from_fields(tpl_row, data, sub_id):
    """根据模板字段顺序生成文件名：字段1_字段2_..._subid.docx。"""
    try:
        fields = json.loads(tpl_row['fields_json'])
        field_names = [f['name'] for f in fields if isinstance(f, dict) and f.get('name')]
    except Exception:
        field_names = []

    if not field_names:
        return f"{safe_filename(data.get('姓名', sub_id))}_{sub_id}.docx"

    parts = []
    for fn in field_names:
        val = data.get(fn)
        if val is not None and str(val).strip():
            parts.append(safe_filename(str(val).strip()))
    if not parts:
        parts.append(str(sub_id))
    name = '_'.join(parts)
    if len(name) > 150:
        name = name[:150]
    return f"{name}.docx"
